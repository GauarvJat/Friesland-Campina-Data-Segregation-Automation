"""
Friesland-Campina  —  Data Segregation Tool  (Streamlit / Web version)
─────────────────────────────────────────────────────────────────────
Run locally : streamlit run friesland_campina_streamlit.py
Deploy      : Push repo to GitHub → connect to streamlit.io
               Set the main file to: friesland_campina_streamlit.py
"""

from __future__ import annotations

import io
import zipfile
from pathlib import PurePath

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION  (identical to EXE version — change here to update both)
# ═══════════════════════════════════════════════════════════════════════════════

PLATFORM_CONFIG: dict[str, list[str]] = {
    "Ad Server":    ["CAMPAIGN", "PLACEMENT", "CREATIVE"],
    "Paid Social":  ["CAMPAIGN", "PLACEMENT", "CREATIVE"],
    "Programmatic": ["CAMPAIGN", "PLACEMENT", "PLACEMENTGROUP"],  # CREATIVE = PLACEMENTGROUP
    "Search":       ["CAMPAIGN", "PLACEMENT", "CREATIVE"],
    "YouTube":      ["CAMPAIGN", "PLACEMENT", "PLACEMENTGROUP"],  # CREATIVE = PLACEMENTGROUP
}

PLATFORM_ALIASES: dict[str, str] = {
    "adserver":     "Ad Server",
    "ad server":    "Ad Server",
    "paidsocial":   "Paid Social",
    "paid social":  "Paid Social",
    "programmatic": "Programmatic",
    "search":       "Search",
    "youtube":      "YouTube",
}

FREE_FORM_COLS: dict[tuple[str, str], list[str]] = {
    ("Ad Server",    "CAMPAIGN"):        ["Campaign Name Description", "Start Date", "End Date", "Free Form Field"],
    ("Ad Server",    "PLACEMENT"):       ["Ad Group or Set Type", "Free Form Field"],
    ("Ad Server",    "CREATIVE"):        ["Ad or Creative Name", "Creative Description", "Free Form Field"],
    ("Programmatic", "CAMPAIGN"):        ["Campaign Name Description", "Start Date", "End Date", "Free Form Field"],
    ("Programmatic", "PLACEMENT"):       ["Ad Group or Set Name", "Free Form Field"],
    ("Programmatic", "PLACEMENTGROUP"):  ["Campaign Name Description", "Start Date", "End Date", "Free Form Field"],
    ("YouTube",      "CAMPAIGN"):        ["Campaign Name Description", "Start Date", "End Date", "Free Form Field"],
    ("YouTube",      "PLACEMENT"):       ["Ad Group or Set Type", "Free Form Field"],
    ("YouTube",      "PLACEMENTGROUP"):  ["Campaign Name Description", "Start Date", "End Date", "Free Form Field"],
    ("Paid Social",  "CAMPAIGN"):        ["Campaign Name Description", "Start Date", "End Date", "Free Form Field"],
    ("Paid Social",  "PLACEMENT"):       ["Ad Group or Set Type", "Free Form Field"],
    ("Paid Social",  "CREATIVE"):        ["Ad or Creative Name", "Creative Description", "Free Form Field"],
    ("Search",       "CAMPAIGN"):        ["Campaign Name Description", "Start Date", "End Date", "Free Form Field"],
    ("Search",       "PLACEMENT"):       ["Ad Group or Set Type", "Free Form Field"],
    ("Search",       "CREATIVE"):        ["Ad or Creative Name"],
}

NAMED_COLS_TO_DROP: list[str] = ["accuracy flag", "status", "clicks", "currency", "spend"]
KNOWN_LEVELS:       list[str] = ["PLACEMENTGROUP", "CAMPAIGN", "PLACEMENT", "CREATIVE"]


# ═══════════════════════════════════════════════════════════════════════════════
#  FILENAME PARSER
# ═══════════════════════════════════════════════════════════════════════════════

def parse_filename(filename: str) -> tuple[str | None, str | None, str | None]:
    """
    Parses filenames of the form:  <Channel> <LEVEL> <Date>.xlsx
    e.g.  AdServer CAMPAIGN 06-04-2026.xlsx
          Paid Social PLACEMENT 06-04-2026.xlsx

    Returns (platform_raw, level, date) or (None, None, None) on failure.
    """
    stem = PurePath(filename).stem.strip()
    parts = stem.split()

    if len(parts) < 3:
        return None, None, None

    for i, token in enumerate(parts):
        if token.upper() in KNOWN_LEVELS:
            if i == 0 or i == len(parts) - 1:
                continue
            return (
                " ".join(parts[:i]).strip(),
                token.upper(),
                " ".join(parts[i + 1:]).strip(),
            )

    return None, None, None


def resolve_platform(platform_raw: str) -> str | None:
    """Maps a raw platform alias to a canonical platform name, or None."""
    return PLATFORM_ALIASES.get(platform_raw.lower())


# ═══════════════════════════════════════════════════════════════════════════════
#  CORE PROCESSING
# ═══════════════════════════════════════════════════════════════════════════════

def process_dataframe(df: pd.DataFrame, platform: str, level: str) -> pd.DataFrame:
    """
    1. Fill every empty cell with 'Is missing' — SKIP Free-Form columns.
    2. Rename duplicate Market column (Market_MK).
    3. Drop all columns whose name contains 'NC'.
    4. Drop named columns: Accuracy Flag, Status, Clicks, Currency, Spend.
    """
    free_form_lower = {f.lower() for f in FREE_FORM_COLS.get((platform, level), [])}

    for col in df.columns:
        if str(col).lower() in free_form_lower:
            continue
        df[col] = df[col].replace("", pd.NA)
        df[col] = df[col].fillna("Is missing")

    market_cols = [c for c in df.columns if str(c).lower() == "market"]
    if len(market_cols) > 1:
        df.rename(columns={market_cols[1]: "Market_MK"}, inplace=True)

    nc_cols = [c for c in df.columns if "NC" in str(c)]
    df.drop(columns=nc_cols, inplace=True, errors="ignore")

    for target in NAMED_COLS_TO_DROP:
        lower_map = {str(c).lower(): c for c in df.columns}
        if target in lower_map:
            df.drop(columns=[lower_map[target]], inplace=True, errors="ignore")

    return df


def apply_formatting_to_workbook(wb, free_form_map: dict):
    """
    Red-highlight every cell containing 'Is missing' or 'Invalid',
    EXCEPT cells in Free-Form columns.
    """
    red_fill   = PatternFill(start_color="FFFF0000", fill_type="solid")
    bad_values = {"is missing", "invalid", "invalid value"}

    for sheet_name in wb.sheetnames:
        ws       = wb[sheet_name]
        ff_lower = {f.lower() for f in free_form_map.get(sheet_name, [])}

        ff_col_indices: set[int] = set()
        for cell in ws[1]:
            if cell.value and str(cell.value).lower() in ff_lower:
                ff_col_indices.add(cell.column)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column in ff_col_indices:
                    continue
                if cell.value and str(cell.value).strip().lower() in bad_values:
                    cell.fill = red_fill

    return wb


# ═══════════════════════════════════════════════════════════════════════════════
#  ORCHESTRATION  (in-memory — no disk I/O needed for Streamlit)
# ═══════════════════════════════════════════════════════════════════════════════

def run_segregation(
    uploaded_files: list,
    progress_cb=None,
) -> tuple[dict[str, bytes], dict]:
    """
    Process a list of Streamlit UploadedFile objects entirely in memory.

    Returns:
        output_files  — { output_filename: xlsx_bytes }
        stats         — summary + skipped list
    """
    stats: dict = {
        "files_created": 0,
        "tabs_created":  0,
        "platforms_processed": 0,
        "skipped": [],
    }
    inputs: dict = {}

    # ── Phase 1: group uploaded files ────────────────────────────────────────
    for uf in uploaded_files:
        platform_raw, level, date = parse_filename(uf.name)
        if not platform_raw:
            stats["skipped"].append(f"{uf.name}  — filename not parseable")
            continue
        platform = resolve_platform(platform_raw)
        if not platform:
            stats["skipped"].append(f"{uf.name}  — unknown channel: '{platform_raw}'")
            continue
        if level not in PLATFORM_CONFIG[platform]:
            stats["skipped"].append(f"{uf.name}  — '{level}' not valid for {platform}")
            continue
        inputs.setdefault(platform, {}).setdefault(date, {})[level] = uf

    output_files: dict[str, bytes] = {}
    total_combos = sum(len(d) for d in inputs.values())
    combo_idx    = 0

    # ── Phase 2: process each platform → date group ──────────────────────────
    for platform, dates in inputs.items():
        stats["platforms_processed"] += 1

        for date, levels in dates.items():
            combo_idx += 1
            if progress_cb:
                progress_cb(combo_idx / max(total_combos, 1))

            market_data: dict = {}

            for level, uf in levels.items():
                try:
                    uf.seek(0)
                    df = pd.read_excel(
                        io.BytesIO(uf.read()),
                        keep_default_na=False,
                        na_values=[""],
                    )
                    if df.empty or "Market" not in df.columns:
                        stats["skipped"].append(
                            f"{uf.name}  — empty or missing 'Market' column"
                        )
                        continue
                    df = process_dataframe(df, platform, level)
                    for market, group in df.groupby("Market"):
                        market_data.setdefault(str(market), {})[level] = group
                except Exception as exc:
                    stats["skipped"].append(f"{uf.name}  — error: {exc}")

            for market, levels_dict in market_data.items():
                safe_market     = str(market).replace("/", "-").replace("\\", "-")
                output_filename = f"{platform}_{safe_market}_{date}.xlsx"
                free_form_map:  dict = {}

                buf = io.BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                    for level in PLATFORM_CONFIG[platform]:
                        if level in levels_dict:
                            levels_dict[level].to_excel(
                                writer, sheet_name=level, index=False
                            )
                            stats["tabs_created"] += 1
                            free_form_map[level] = FREE_FORM_COLS.get(
                                (platform, level), []
                            )

                buf.seek(0)
                wb = load_workbook(buf)
                apply_formatting_to_workbook(wb, free_form_map)

                final = io.BytesIO()
                wb.save(final)
                output_files[output_filename] = final.getvalue()
                stats["files_created"] += 1

    return output_files, stats


def build_zip(output_files: dict[str, bytes]) -> bytes:
    """Pack all output xlsx files into one downloadable ZIP."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for fname, data in output_files.items():
            zf.writestr(fname, data)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ═══════════════════════════════════════════════════════════════════════════════

# ── Page config (must be FIRST Streamlit call) ───────────────────────────────
st.set_page_config(
    page_title = "FrieslandCampina — Data Segregation",
    page_icon  = "🧀",
    layout     = "centered",
)

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ── Global ── */
html, body, [data-testid="stAppViewContainer"] {
    background-color: #0B1F3A !important;
}
[data-testid="stAppViewContainer"] > .main {
    background-color: #0B1F3A;
}
[data-testid="stHeader"]          { background: transparent !important; }
[data-testid="stToolbar"]         { display: none; }
footer                            { visibility: hidden; }

/* ── Hero banner ── */
.fc-hero {
    background: linear-gradient(135deg, #0E2A50 0%, #132E58 100%);
    border-bottom: 4px solid #F7A800;
    border-radius: 12px;
    padding: 32px 36px 26px;
    margin-bottom: 28px;
}
.fc-hero-name {
    font-size: 2rem; font-weight: 900;
    color: #F7A800; letter-spacing: .5px;
    margin: 0 0 4px;
}
.fc-hero-sub {
    font-size: .95rem; color: #8AAFC8; margin: 0 0 18px;
}
.fc-tags { display: flex; flex-wrap: wrap; gap: 8px; }
.fc-tag {
    background: rgba(255,255,255,.06);
    border: 1px solid #1A3A5C;
    color: #8AAFC8; font-size: .78rem; font-weight: 600;
    padding: 3px 12px; border-radius: 20px;
}

/* ── Info / expander card ── */
.fc-info {
    background: #112340;
    border: 1px solid #1A3A5C;
    border-radius: 10px;
    padding: 16px 20px;
    margin-bottom: 18px;
    font-size: .88rem; color: #8AAFC8;
}
.fc-info code {
    background: #07111F; color: #7DD4C0;
    padding: 1px 6px; border-radius: 4px;
    font-size: .85rem;
}

/* ── Section labels ── */
.fc-label {
    font-size: .72rem; font-weight: 800;
    letter-spacing: 1.5px; text-transform: uppercase;
    color: #8AAFC8; margin-bottom: 6px;
}

/* ── Upload zone override ── */
[data-testid="stFileUploader"] section {
    background: #112340 !important;
    border: 1.5px dashed #1E4070 !important;
    border-radius: 10px !important;
}
[data-testid="stFileUploader"] section:hover {
    border-color: #F7A800 !important;
}
[data-testid="stFileUploader"] label { color: #8AAFC8 !important; }

/* ── Buttons ── */
.stButton > button {
    background: #0075BE !important;
    color: #fff !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 800 !important;
    font-size: 1rem !important;
    padding: 12px 0 !important;
    width: 100% !important;
    letter-spacing: .3px;
    transition: background .2s;
}
.stButton > button:hover { background: #005A9A !important; }
.stButton > button:disabled { opacity: .5 !important; }

/* ── Download button ── */
.stDownloadButton > button {
    background: #F7A800 !important;
    color: #0B1F3A !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 800 !important;
    font-size: .95rem !important;
    padding: 12px 0 !important;
    width: 100% !important;
}
.stDownloadButton > button:hover { background: #D99200 !important; }

/* ── Metric boxes ── */
[data-testid="metric-container"] {
    background: #112340;
    border: 1px solid #1A3A5C;
    border-radius: 10px;
    padding: 14px !important;
}
[data-testid="metric-container"] label  { color: #8AAFC8 !important; font-size: .8rem !important; }
[data-testid="stMetricValue"]           { color: #F7A800 !important; }

/* ── Alerts ── */
[data-testid="stAlert"] {
    background: #112340 !important;
    border: 1px solid #1A3A5C !important;
    border-radius: 8px !important;
    color: #8AAFC8 !important;
}

/* ── Progress bar ── */
[data-testid="stProgressBar"] > div > div { background: #F7A800 !important; }

/* ── Expander ── */
[data-testid="stExpander"] {
    background: #112340 !important;
    border: 1px solid #1A3A5C !important;
    border-radius: 10px !important;
}
.streamlit-expanderHeader { color: #8AAFC8 !important; font-size: .9rem !important; }

/* ── Divider ── */
hr { border-color: #1A3A5C !important; }

/* ── General text ── */
p, li, span { color: #8AAFC8; }
strong      { color: #E2EBF5 !important; }
code        { background: #07111F !important; color: #7DD4C0 !important;
              border-radius: 4px !important; font-size: .85rem !important; }

/* ── Footer ── */
.fc-footer {
    text-align: center; color: #1A3A5C;
    font-size: .78rem; padding: 24px 0 8px;
    border-top: 1px solid #1A3A5C; margin-top: 40px;
}
</style>
""", unsafe_allow_html=True)


# ── Hero ─────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="fc-hero">
    <div class="fc-hero-name">🧀 FrieslandCampina</div>
    <div class="fc-hero-sub">Data Segregation Tool &nbsp;·&nbsp; v2.0</div>
    <div class="fc-tags">
        <span class="fc-tag">● Ad Server</span>
        <span class="fc-tag">● Paid Social</span>
        <span class="fc-tag">● Programmatic</span>
        <span class="fc-tag">● Search</span>
        <span class="fc-tag">● YouTube</span>
    </div>
</div>
""", unsafe_allow_html=True)


# ── File naming guide ─────────────────────────────────────────────────────────
with st.expander("📋  File Naming Convention — click to expand"):
    st.markdown("""
Each uploaded file **must** follow this exact pattern:

```
<Channel> <LEVEL> <Date>.xlsx
```

**Valid examples:**
```
AdServer CAMPAIGN 06-04-2026.xlsx
Paid Social PLACEMENT 06-04-2026.xlsx
Programmatic PLACEMENTGROUP 06-04-2026.xlsx
YouTube PLACEMENTGROUP 06-04-2026.xlsx
Search CREATIVE 06-04-2026.xlsx
```

| Channel | Valid Levels |
|---|---|
| Ad Server | CAMPAIGN · PLACEMENT · CREATIVE |
| Paid Social | CAMPAIGN · PLACEMENT · CREATIVE |
| Search | CAMPAIGN · PLACEMENT · CREATIVE |
| Programmatic | CAMPAIGN · PLACEMENT · PLACEMENTGROUP |
| YouTube | CAMPAIGN · PLACEMENT · PLACEMENTGROUP |

> **Note:** For Programmatic and YouTube, `CREATIVE` is equivalent to `PLACEMENTGROUP`.
""")


# ── Upload section ────────────────────────────────────────────────────────────
st.markdown('<div class="fc-label">Upload Input Files</div>', unsafe_allow_html=True)

uploaded_files = st.file_uploader(
    label           = "Drop your .xlsx files here or click to browse",
    type            = ["xlsx"],
    accept_multiple_files = True,
    label_visibility= "collapsed",
)

if uploaded_files:
    st.info(f"**{len(uploaded_files)}** file(s) selected and ready to process.")

st.markdown("<br>", unsafe_allow_html=True)


# ── Run button ────────────────────────────────────────────────────────────────
run_clicked = st.button(
    "▶   Run Segregation",
    disabled = not bool(uploaded_files),
)

# ── Processing ────────────────────────────────────────────────────────────────
if run_clicked and uploaded_files:
    progress_bar = st.progress(0, text="Initialising…")

    def update_progress(fraction: float):
        pct = min(int(fraction * 100), 99)
        progress_bar.progress(pct, text=f"Processing…  {pct}%")

    with st.spinner("Running segregation — please wait…"):
        try:
            output_files, stats = run_segregation(
                uploaded_files,
                progress_cb = update_progress,
            )
        except Exception as exc:
            st.error(f"Fatal error: {exc}")
            st.stop()

    progress_bar.progress(100, text="Done ✅")
    st.markdown("---")

    # ── Stats ────────────────────────────────────────────────────────────────
    st.markdown("### ✅ Segregation Complete")
    c1, c2, c3 = st.columns(3)
    c1.metric("Platforms processed",  stats["platforms_processed"])
    c2.metric("Output files created", stats["files_created"])
    c3.metric("Tabs written",         stats["tabs_created"])

    # ── Skipped files ─────────────────────────────────────────────────────────
    if stats["skipped"]:
        with st.expander(f"⚠️  {len(stats['skipped'])} file(s) skipped", expanded=True):
            for msg in stats["skipped"]:
                st.markdown(f"- `{msg}`")

    # ── Download ──────────────────────────────────────────────────────────────
    if output_files:
        st.markdown("---")
        st.markdown('<div class="fc-label">Download Output</div>', unsafe_allow_html=True)

        if len(output_files) == 1:
            fname, data = next(iter(output_files.items()))
            st.download_button(
                label     = f"⬇️  Download  {fname}",
                data      = data,
                file_name = fname,
                mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            zip_bytes = build_zip(output_files)
            st.download_button(
                label     = f"⬇️  Download All  ({len(output_files)} files)  as ZIP",
                data      = zip_bytes,
                file_name = "FrieslandCampina_Segregated_Output.zip",
                mime      = "application/zip",
            )
            with st.expander(f"📂  Files in the ZIP  ({len(output_files)})"):
                for fname in sorted(output_files.keys()):
                    st.markdown(f"- `{fname}`")
    else:
        st.warning(
            "No output files were generated. "
            "Check your filenames match the required pattern above."
        )


# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown(
    '<div class="fc-footer">© FrieslandCampina &nbsp;·&nbsp; Internal Use Only</div>',
    unsafe_allow_html=True,
)
